#!/usr/bin/env python3
"""
Excel 生成脚本
根据提取数据 JSON 生成汇总 Excel 报表
遵循 xlsx-format.md 格式规范

输出路径: output/doc_scan_result_{timestamp}.xlsx
"""
import json
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 输出目录
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

# 样式常量（遵循 xlsx-format.md 规范）
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

# 分类组映射（基于 document_classification.md V2.0，10 大分类组、37 个子类型）
GROUP_MAP = {
    # A. 财务与税务类 (finance-tax)
    "增值税发票": "A.财务与税务",
    "银行回单": "A.财务与税务",
    "银行对账单": "A.财务与税务",
    "费用报销单": "A.财务与税务",
    "合同协议": "A.财务与税务",
    "税务申报表": "A.财务与税务",
    "财务报表": "A.财务与税务",
    "收据付款凭证": "A.财务与税务",
    
    # B. 人力资源类 (hr)
    "简历": "B.人力资源",
    "劳动合同": "B.人力资源",
    "离职证明": "B.人力资源",
    "工资条": "B.人力资源",
    "考勤记录": "B.人力资源",
    "培训证书": "B.人力资源",
    "绩效考核表": "B.人力资源",
    
    # C. 供应链与采购类 (supply-chain)
    "采购订单": "C.供应链与采购",
    "送货单": "C.供应链与采购",
    "入库单": "C.供应链与采购",
    "质检报告": "C.供应链与采购",
    "供应商评估表": "C.供应链与采购",
    "库存盘点表": "C.供应链与采购",
    
    # D. 行政与法务类 (admin-legal)
    "营业执照": "D.行政与法务",
    "身份证": "D.行政与法务",
    "护照": "D.行政与法务",
    "保密协议": "D.行政与法务",
    "资质证书": "D.行政与法务",
    "公文通知": "D.行政与法务",
    
    # E. 医疗类 (medical)
    "病历": "E.医疗",
    "处方单": "E.医疗",
    "检验检查报告": "E.医疗",
    "体检报告": "E.医疗",
    
    # F. 保险类 (insurance)
    "保单": "F.保险",
    "理赔申请": "F.保险",
    "理赔结案通知": "F.保险",
    
    # G. 物流类 (logistics)
    "运单": "G.物流",
    "提单": "G.物流",
    "报关单": "G.物流",
    
    # H. 技术与运维类 (tech-ops)
    "系统日志": "H.技术与运维",
    "漏洞扫描报告": "H.技术与运维",
    "服务器监控报表": "H.技术与运维",
    
    # I. 客服与销售类 (sales-service)
    "客诉工单记录": "I.客服与销售",
    "销售报价单": "I.客服与销售",
    "售后退换货单": "I.客服与销售",
    
    # J. 政务与合规类 (gov-compliance)
    "招投标文件": "J.政务与合规",
    "政务审批单": "J.政务与合规",
    "合规审计报告": "J.政务与合规",
    
    # 其他
    "未识别": "未识别",
    "解析失败": "解析失败",
}


def display_width(s: str) -> int:
    """计算字符串显示宽度（CJK 字符宽度为 2）"""
    if not s:
        return 0
    return sum(2 if ord(c) > 0x7F else 1 for c in str(s))


def auto_column_width(ws, headers: list, rows: list, max_width: int = 50):
    """自动调整列宽"""
    for col_idx in range(1, len(headers) + 1):
        widths = [display_width(headers[col_idx - 1])]
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1]:
                widths.append(display_width(str(row[col_idx - 1])))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(widths) + 2, max_width)


def create_data_sheet(wb: Workbook, sheet_name: str, headers: list, rows: list):
    """创建数据 Sheet"""
    ws = wb.create_sheet(title=sheet_name[:31])
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 写入数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
    
    # 自动列宽
    auto_column_width(ws, headers, rows)
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 自动筛选（仅有数据时）
    if rows:
        ws.auto_filter.ref = ws.dimensions


def create_summary_sheet(wb: Workbook, extraction_data: dict, scan_time: str, total_files: int):
    """创建汇总 Sheet（放在首位）"""
    ws = wb.create_sheet(title="汇总", index=0)
    
    headers = ["分类组", "子类型", "文件数量", "提取字段"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    row_idx = 2
    for subtype_name, data in extraction_data.items():
        group_name = GROUP_MAP.get(subtype_name, "其他")
        headers_cn = data.get("headers_cn", [])
        fields = ", ".join(headers_cn[1:]) if len(headers_cn) > 1 else ""
        
        ws.cell(row_idx, 1, value=group_name).border = THIN_BORDER
        ws.cell(row_idx, 2, value=subtype_name).border = THIN_BORDER
        ws.cell(row_idx, 3, value=len(data.get("rows", []))).border = THIN_BORDER
        ws.cell(row_idx, 4, value=fields).border = THIN_BORDER
        row_idx += 1
    
    # 元信息
    row_idx += 1
    ws.cell(row_idx, 1, value="扫描时间").font = Font(bold=True)
    ws.cell(row_idx, 2, value=scan_time)
    row_idx += 1
    ws.cell(row_idx, 1, value="文件总数").font = Font(bold=True)
    ws.cell(row_idx, 2, value=total_files)
    
    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60


def generate_excel(input_json_path: str, output_path: str = None) -> str:
    """
    根据提取数据 JSON 生成 Excel
    
    JSON 格式：
    {
        "scan_time": "2026-03-30 10:00:00",
        "total_files": 10,
        "extraction_data": {
            "增值税发票": {
                "headers_cn": ["源文件名", "发票类型", "发票代码", ...],
                "rows": [["发票_001.pdf", "专用", "1234567890", ...], ...]
            },
            ...
        }
    }
    
    Args:
        input_json_path: 输入 JSON 文件路径
        output_path: 输出 Excel 路径（默认 output/doc_scan_result_{timestamp}.xlsx）
    
    Returns:
        输出 Excel 文件路径
    """
    with open(input_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    extraction_data = data.get("extraction_data", {})
    scan_time = data.get("scan_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    total_files = data.get("total_files", sum(len(d.get("rows", [])) for d in extraction_data.values()))
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 为每个子类型创建 Sheet
    for subtype_name, subtype_data in extraction_data.items():
        headers = subtype_data.get("headers_cn", [])
        rows = subtype_data.get("rows", [])
        if headers:
            create_data_sheet(wb, subtype_name, headers, rows)
    
    # 创建汇总 Sheet
    create_summary_sheet(wb, extraction_data, scan_time, total_files)
    
    # 输出路径（默认 output/doc_scan_result_{timestamp}.xlsx）
    if output_path:
        output_path = Path(output_path).resolve()
    else:
        # 默认输出到 output 目录
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"doc_scan_result_{timestamp}.xlsx"
    
    wb.save(str(output_path))
    print(f"[保存] ✓ Excel 结果已保存到: {output_path}", flush=True)
    return str(output_path)


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="根据提取数据 JSON 生成汇总 Excel 报表"
    )
    parser.add_argument(
        "input_path",
        help="输入 JSON 文件路径"
    )
    parser.add_argument(
        "output_path",
        nargs="?",
        default=None,
        help="输出 Excel 路径（可选，默认 output/doc_scan_result_{timestamp}.xlsx）"
    )
    parser.add_argument(
        "--workspace-dir",
        default=None,
        help="用户工作目录路径"
    )

    args = parser.parse_args()

    # 设置工作目录（必须在任何 read_config() 之前）
    if args.workspace_dir:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from common.config_loader import set_workspace_dir
        set_workspace_dir(args.workspace_dir)

    if not Path(args.input_path).exists():
        print(json.dumps({"error": f"文件不存在: {args.input_path}"}, ensure_ascii=False))
        sys.exit(1)

    try:
        result_path = generate_excel(args.input_path, args.output_path)
        print(json.dumps({
            "status": "success",
            "output_path": result_path
        }, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
