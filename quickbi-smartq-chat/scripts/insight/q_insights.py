# -*- coding: utf-8 -*-
"""
Quick BI 小Q解读：Excel 文件解析 / 仪表板快照 + 数据解读流式输出。

用法：
    # 通过 Excel 文件解读
    python scripts/q_insights.py "这个报表有什么异常？" --excel-file "/path/to/data.xlsx"

    # 通过仪表板快照解读（旧模式）
    python scripts/q_insights.py "这个报表有什么异常？" --works-id "your-works-id"
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

from common.utils import (
    read_config,
    require_user_id,
    request_openapi,
    request_openapi_stream,
    parse_sse_event,
    set_workspace_dir,
)

SNAPSHOT_URI = "/openapi/v2/snapshot/calling/shot"
INTERPRETATION_URI = "/openapi/v2/smartq/dataInterpretationStream"

POLL_INTERVAL_SECONDS = 3
MAX_POLL_COUNT = 60
MAX_MARKDOWN_CHARS = 100000

ACC_PROMPT = """【关键要求】：若输入数据中未包含用户提问的相关内容，则直接回复“不存在”，禁止随意捏造数据。若报告中包含相关数据，且用户问题有明确指向，则须在报告开头设置“用户问题解答”章节，针对问题作出直接、明确的回答，并予以标注。"""


# ---------------------------------------------------------------------------
# Excel 文件解析
# ---------------------------------------------------------------------------

def _parse_xls_to_markdown(file_path: str) -> Optional[str]:
    """解析 .xls 文件（旧版 Excel 97-2003 格式），需要 xlrd 库。"""
    try:
        import xlrd
    except ImportError:
        print("[Excel] 缺少 xlrd 依赖，请执行: pip install xlrd", flush=True)
        return None

    try:
        wb = xlrd.open_workbook(file_path)
    except Exception as e:
        print(f"[Excel] 文件打开失败: {e}", flush=True)
        return None

    md_parts: List[str] = []

    for sheet in wb.sheets():
        if sheet.nrows == 0:
            continue

        rows: List[List[str]] = []
        for row_idx in range(sheet.nrows):
            rows.append([str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)])

        headers = rows[0]
        col_count = len(headers)

        lines: List[str] = []
        lines.append(f"## {sheet.name}\n")
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join([" --- "] * col_count) + "|")

        for data_row in rows[1:]:
            padded = data_row + [""] * (col_count - len(data_row))
            lines.append("| " + " | ".join(padded[:col_count]) + " |")

        md_parts.append("\n".join(lines))

    if not md_parts:
        print("[Excel] 文件中无有效数据", flush=True)
        return None

    markdown_text = "\n\n".join(md_parts)
    print(f"[Excel] 解析完成，共 {wb.nsheets} 个 Sheet，数据长度: {len(markdown_text)} 字符", flush=True)
    return markdown_text


def _parse_xlsx_to_markdown(file_path: str) -> Optional[str]:
    """解析 .xlsx 文件（Office Open XML 格式），需要 openpyxl 库。"""
    try:
        import openpyxl
    except ImportError:
        print("[Excel] 缺少 openpyxl 依赖，请执行: pip install openpyxl", flush=True)
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[Excel] 文件打开失败: {e}", flush=True)
        return None

    md_parts: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            continue

        headers = rows[0]
        col_count = len(headers)

        lines: List[str] = []
        lines.append(f"## {sheet_name}\n")
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join([" --- "] * col_count) + "|")

        for data_row in rows[1:]:
            padded = data_row + [""] * (col_count - len(data_row))
            lines.append("| " + " | ".join(padded[:col_count]) + " |")

        md_parts.append("\n".join(lines))

    wb.close()

    if not md_parts:
        print("[Excel] 文件中无有效数据", flush=True)
        return None

    markdown_text = "\n\n".join(md_parts)
    print(f"[Excel] 解析完成，共 {len(wb.sheetnames)} 个 Sheet，数据长度: {len(markdown_text)} 字符", flush=True)
    return markdown_text


def parse_excel_to_markdown(file_path: str) -> Optional[str]:
    """
    将 Excel 文件解析为 Markdown 表格文本。

    自动识别 .xls / .xlsx 格式，支持多 Sheet。
    返回合并后的 Markdown 文本；解析失败时返回 None。
    """
    if not os.path.isfile(file_path):
        print(f"[Excel] 文件不存在: {file_path}", flush=True)
        return None

    print(f"[Excel] 正在解析文件: {file_path}", flush=True)

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        return _parse_xls_to_markdown(file_path)
    elif ext == ".xlsx":
        return _parse_xlsx_to_markdown(file_path)
    else:
        print(f"[Excel] 不支持的文件格式: {ext}，仅支持 .xls 和 .xlsx", flush=True)
        return None


# ---------------------------------------------------------------------------
# 仪表板快照
# ---------------------------------------------------------------------------

def call_snapshot(
    works_id: str,
    user_id: str,
    *,
    config: Optional[dict] = None,
) -> dict:
    """
    调用 POST /openapi/v2/snapshot/calling/shot 拉取仪表板快照。

    返回 SnapshotResultModel：
      {status: "processing"|"success"|"failed", errorInfo: {...}, resultMarkdownText: "..."}
    """
    payload: Dict[str, Any] = {
        "worksId": works_id,
        "worksType": "dashboard",
        "userId": user_id,
        "targetType": "excel",
    }
    resp = request_openapi(
        "POST",
        SNAPSHOT_URI,
        json_body=payload,
        timeout=60,
        config=config,
    )
    return resp.json()


def poll_snapshot(
    works_id: str,
    user_id: str,
    *,
    config: Optional[dict] = None,
) -> Optional[str]:
    """
    轮询快照接口直到 status 为 success 或 failed。

    返回 resultMarkdownText（成功时）或 None（失败 / 超时时）。
    """
    for attempt in range(1, MAX_POLL_COUNT + 1):
        result = call_snapshot(works_id, user_id, config=config)

        status = result.get("status", "")
        print(f"[快照] 第 {attempt} 次轮询，状态: {status}", flush=True)

        if status == "success":
            markdown_text = result.get("resultMarkdownText", "")
            if not markdown_text:
                print("[快照] 状态为 success 但 resultMarkdownText 为空", flush=True)
                return None
            print(f"[快照] 快照数据获取成功，数据长度: {len(markdown_text)} 字符", flush=True)
            return markdown_text

        if status == "failed":
            error_info = result.get("errorInfo", {})
            print(f"[快照] 仪表板数据处理失败: {json.dumps(error_info, ensure_ascii=False)}", flush=True)
            print("[快照] 仪表板数据处理失败，请联系产品服务同学处理", flush=True)
            return None

        if status == "processing":
            time.sleep(POLL_INTERVAL_SECONDS)
            continue

        print(f"[快照] 未知状态: {status}，原始响应: {json.dumps(result, ensure_ascii=False)}", flush=True)
        return None

    print(f"[快照] 轮询超时（已等待 {MAX_POLL_COUNT * POLL_INTERVAL_SECONDS} 秒）", flush=True)
    return None


# ---------------------------------------------------------------------------
# 数据解读（SSE 流式）
# ---------------------------------------------------------------------------

def run_interpretation_stream(
    string_data: str,
    user_question: str,
    *,
    config: Optional[dict] = None,
):
    """
    调用 POST /openapi/v2/smartq/dataInterpretationStream 进行数据解读。

    实时解析 SSE 事件并输出推理过程和解读结果。
    """
    config = config or read_config()
    oapi_user_id = require_user_id(config)

    payload: Dict[str, Any] = {
        "stringData": base64.b64encode(string_data.encode("utf-8")).decode("utf-8"),
        "userQuestion": ACC_PROMPT + user_question,
        "modelCode": config.get("model_code", "SYSTEM_deepseek-v3"),
        "oapiUserId": oapi_user_id,
    }

    print(f"\n{'=' * 60}", flush=True)
    print(f"[数据解读] 问题: {user_question}", flush=True)
    print(f"{'=' * 60}\n", flush=True)

    reasoning_buf: List[str] = []
    text_buf: List[str] = []
    event_count = 0

    for raw_event in request_openapi_stream(
        INTERPRETATION_URI, json_body=payload, config=config, timeout=600
    ):
        event_count += 1
        # print(f"[DEBUG] 原始事件 #{event_count}: {raw_event[:300]}", flush=True)

        event_data = parse_sse_event(raw_event)
        if not event_data:
            print(f"[DEBUG] 事件 #{event_count} 解析为空，跳过", flush=True)
            continue

        event_type = event_data.get("type", "")
        data = event_data.get("data", "")
        # print(f"[DEBUG] 事件 #{event_count}: type={event_type}, data={str(data)[:200]}", flush=True)

        if event_type in ("heartbeat", "trace", "locale"):
            continue

        elif event_type == "reasoning":
            reasoning_buf.append(str(data))

        elif event_type in ("text", "summary"):
            text_buf.append(str(data))

        elif event_type == "finish":
            break

        else:
            print(f"[SSE] 未处理事件: type={event_type}, data={json.dumps(event_data, ensure_ascii=False)[:500]}", flush=True)

    print(f"[DEBUG] SSE 流结束，共收到 {event_count} 个事件", flush=True)

    if reasoning_buf:
        reasoning_text = "".join(reasoning_buf)
        print(f"[推理过程]\n{reasoning_text}\n", flush=True)

    if text_buf:
        interpretation_text = "".join(text_buf)
        print(f"[解读结果]\n{interpretation_text}", flush=True)
    else:
        print("[解读结果] 未获取到解读内容", flush=True)

    print(f"\n[完成] 数据解读结束", flush=True)
    return "".join(text_buf)


# ---------------------------------------------------------------------------
# 完整流程
# ---------------------------------------------------------------------------

def run_insights(
    question: str,
    works_id: Optional[str] = None,
    *,
    excel_file: Optional[str] = None,
    config: Optional[dict] = None,
):
    """执行小Q解读的完整流程：Excel 解析 / 快照轮询 → 数据解读。"""
    config = config or read_config()

    # 统一调用 require_user_id，确保试用注册流程正常执行
    user_id = require_user_id(config)

    print(f"{'=' * 60}", flush=True)
    print(f"[小Q解读] 问题: {question}", flush=True)

    # 优先从 Excel 文件获取数据
    if excel_file:
        print(f"[小Q解读] 数据来源: Excel 文件 ({excel_file})", flush=True)
        print(f"{'=' * 60}\n", flush=True)
        markdown_text = parse_excel_to_markdown(excel_file)
    elif works_id:
        print(f"[小Q解读] 数据来源: 仪表板快照 ({works_id})", flush=True)
        print(f"{'=' * 60}\n", flush=True)
        print("[快照] 开始拉取仪表板快照数据...", flush=True)
        markdown_text = poll_snapshot(works_id, user_id, config=config)
    else:
        print(f"{'=' * 60}\n", flush=True)
        print("[小Q解读] 流程终止：必须指定 --excel-file 或 --works-id", flush=True)
        sys.exit(1)

    if not markdown_text:
        print(f"\n{'=' * 60}", flush=True)
        print("[小Q解读] 流程终止：无法获取数据", flush=True)
        print(f"{'=' * 60}", flush=True)
        sys.exit(1)

    # 数据超限检查：超限时报错终止，要求 Agent 先做数据过滤
    if len(markdown_text) > MAX_MARKDOWN_CHARS:
        print(f"[小Q解读] 数据量超限：当前 {len(markdown_text)} 字符，上限 {MAX_MARKDOWN_CHARS} 字符。", flush=True)
        print("[小Q解读] 请先根据用户问题对 Excel 数据进行过滤（只保留相关行和列），将结果另存为新文件后重新调用。", flush=True)
        print("[小Q解读] 若过滤后仍超限，请按行拆分为多份文件分批调用，最后汇总结果。", flush=True)
        sys.exit(1)

    # Step 3 & 4: 调用数据解读流式接口
    result = run_interpretation_stream(
        string_data=markdown_text,
        user_question=question,
        config=config
    )

    print(f"\n{'=' * 60}", flush=True)
    print("[小Q解读] 解读流程完成", flush=True)
    print(f"{'=' * 60}", flush=True)

    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Quick BI 小Q解读")
    parser.add_argument("question", help="用户的解读问题")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--excel-file", dest="excel_file", help="Excel 文件路径（.xlsx）")
    # group.add_argument("--works-id", dest="works_id", help="仪表板 ID (worksId)")
    parser.add_argument("--workspace-dir", default=None, help="用户工作目录路径")
    args = parser.parse_args()

    if args.workspace_dir:
        set_workspace_dir(args.workspace_dir)

    run_insights(args.question, excel_file=args.excel_file)


if __name__ == "__main__":
    main()
