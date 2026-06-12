#!/usr/bin/env python3
"""Spreadsheet formula recalculation via LibreOffice.

Deploys a LibreOffice Basic macro, invokes headless recalculation on an Excel
workbook, then scans all cells for error markers and emits a structured JSON
report to stdout.

Usage:
    python recalc.py <excel_file> [timeout_seconds]
"""

from __future__ import annotations

import json
import platform
import subprocess
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

from office.soffice import get_soffice_env

# ─── Platform Constants ──────────────────────────────────────────────

_SYSTEM = platform.system()

_MACRO_BASE_DIRS: dict[str, str] = {
    "Darwin": "~/Library/Application Support/LibreOffice/4/user/basic/Standard",
    "Linux": "~/.config/libreoffice/4/user/basic/Standard",
}

_MACRO_MODULE_NAME = "Module1.xba"

_MACRO_ENTRY_POINT = (
    "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
    "?language=Basic&location=application"
)

_MACRO_XML = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<!DOCTYPE script:module PUBLIC '
    '"-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">\n'
    '<script:module xmlns:script="http://openoffice.org/2000/script" '
    'script:name="Module1" script:language="StarBasic">\n'
    "    Sub RecalculateAndSave()\n"
    "      ThisComponent.calculateAll()\n"
    "      ThisComponent.store()\n"
    "      ThisComponent.close(True)\n"
    "    End Sub\n"
    "</script:module>"
)

# ─── Excel Error Markers ─────────────────────────────────────────────

CELL_ERROR_MARKERS: tuple[str, ...] = (
    "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A",
)

_MAX_ERROR_LOCATIONS = 20
_DEFAULT_TIMEOUT_SECONDS = 30
_TIMEOUT_EXIT_CODE = 124


# ─── Data Structures ─────────────────────────────────────────────────

@dataclass
class ErrorBucket:
    """Aggregated info for one category of cell errors."""
    count: int
    locations: list[str]


@dataclass
class RecalcReport:
    """Structured result of a recalculation + error-scan pass."""
    status: str = "success"
    total_errors: int = 0
    total_formulas: int = 0
    error_summary: dict[str, ErrorBucket] = field(default_factory=dict)
    error: Optional[str] = None

    def to_dict(self) -> dict:
        """Serialise to a plain dict matching the original JSON contract."""
        result: dict = {}
        if self.error is not None:
            result["error"] = self.error
            return result
        result["status"] = self.status
        result["total_errors"] = self.total_errors
        result["error_summary"] = {
            kind: asdict(bucket) for kind, bucket in self.error_summary.items()
        }
        result["total_formulas"] = self.total_formulas
        return result


# ─── Timeout Utility ─────────────────────────────────────────────────

def _has_gtimeout() -> bool:
    """Check whether GNU coreutils ``gtimeout`` binary exists on this host."""
    try:
        subprocess.run(
            ["gtimeout", "--version"],
            capture_output=True,
            timeout=1,
            check=False,
        )
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False
    return True


# ─── Macro Deployment ────────────────────────────────────────────────

def _deploy_macro() -> bool:
    """Ensure the RecalculateAndSave Basic macro is present in the LO user dir.

    Returns ``True`` on success, ``False`` if LibreOffice is missing or the
    macro file could not be written.
    """
    base_template = _MACRO_BASE_DIRS.get(_SYSTEM, _MACRO_BASE_DIRS["Linux"])
    base_dir = Path(base_template).expanduser()
    target = base_dir / _MACRO_MODULE_NAME

    # Already deployed?
    if target.exists() and "RecalculateAndSave" in target.read_text():
        return True

    # Bootstrap LibreOffice profile when the directory tree is missing.
    if not base_dir.exists():
        try:
            subprocess.run(
                ["soffice", "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=10,
                env=get_soffice_env(),
            )
        except FileNotFoundError:
            return False
        base_dir.mkdir(parents=True, exist_ok=True)

    try:
        target.write_text(_MACRO_XML)
    except OSError:
        return False
    return True


# ─── Workbook Inspection ─────────────────────────────────────────────

def _count_formulas(filepath: Path) -> int:
    """Open *filepath* without cached values and count formula cells."""
    wb = openpyxl.load_workbook(str(filepath), data_only=False)
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total += 1
    wb.close()
    return total


def _scan_errors(filepath: Path) -> tuple[int, dict[str, ErrorBucket]]:
    """Scan cached cell values for Excel error strings.

    Returns a ``(total_error_count, buckets)`` tuple where *buckets* maps each
    error marker that has at least one occurrence to an :class:`ErrorBucket`.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    raw_buckets: dict[str, list[str]] = {m: [] for m in CELL_ERROR_MARKERS}
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                for marker in CELL_ERROR_MARKERS:
                    if marker in cell.value:
                        raw_buckets[marker].append(
                            f"{sheet_name}!{cell.coordinate}"
                        )
                        total += 1
                        break

    wb.close()

    buckets: dict[str, ErrorBucket] = {}
    for kind, locs in raw_buckets.items():
        if locs:
            buckets[kind] = ErrorBucket(
                count=len(locs),
                locations=locs[:_MAX_ERROR_LOCATIONS],
            )

    return total, buckets


# ─── Core Recalculation ──────────────────────────────────────────────

def _build_command(filepath: str, timeout: int) -> list[str]:
    """Assemble the ``soffice`` invocation with optional timeout wrapper."""
    argv = [
        "soffice", "--headless", "--norestore",
        _MACRO_ENTRY_POINT,
        filepath,
    ]

    if _SYSTEM == "Linux":
        argv = ["timeout", str(timeout)] + argv
    elif _SYSTEM == "Darwin" and _has_gtimeout():
        argv = ["gtimeout", str(timeout)] + argv

    return argv


def recalc(filename: str, timeout: int = _DEFAULT_TIMEOUT_SECONDS) -> dict:
    """Run the LO macro on *filename* then inspect the result for errors.

    Returns a plain ``dict`` suitable for JSON serialisation.
    """
    source = Path(filename)

    if not source.exists():
        return RecalcReport(error=f"File {filename} does not exist").to_dict()

    absolute_path = str(source.absolute())

    if not _deploy_macro():
        return RecalcReport(error="Failed to setup LibreOffice macro").to_dict()

    argv = _build_command(absolute_path, timeout)

    try:
        proc = subprocess.run(
            argv, capture_output=True, text=True, env=get_soffice_env(),
        )
    except FileNotFoundError:
        return RecalcReport(
            error="LibreOffice (soffice) is not installed or not in PATH",
        ).to_dict()

    if proc.returncode not in (0, _TIMEOUT_EXIT_CODE):
        msg = proc.stderr or "Unknown error during recalculation"
        if "Module1" in msg or "RecalculateAndSave" not in msg:
            return RecalcReport(
                error="LibreOffice macro not configured properly",
            ).to_dict()
        return RecalcReport(error=msg).to_dict()

    try:
        err_total, buckets = _scan_errors(source)

        report = RecalcReport(
            status="success" if err_total == 0 else "errors_found",
            total_errors=err_total,
            error_summary=buckets,
            total_formulas=_count_formulas(source),
        )
        return report.to_dict()

    except Exception as exc:
        return RecalcReport(error=str(exc)).to_dict()


# ─── CLI Entry Point ─────────────────────────────────────────────────

def main() -> None:
    """Parse CLI arguments and run recalculation."""
    if len(sys.argv) < 2:
        sys.stderr.write(
            "Usage: python recalc.py <excel_file> [timeout_seconds]\n\n"
            "Recalculates all formulas in an Excel file using LibreOffice\n\n"
            "Returns JSON with error details:\n"
            "  - status: 'success' or 'errors_found'\n"
            "  - total_errors: Total number of Excel errors found\n"
            "  - total_formulas: Number of formulas in the file\n"
            "  - error_summary: Breakdown by error type with locations\n"
            "    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A\n"
        )
        sys.exit(1)

    filepath = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else _DEFAULT_TIMEOUT_SECONDS

    print(json.dumps(recalc(filepath, timeout), indent=2))


if __name__ == "__main__":
    main()
